from PyPDF2 import PdfFileReader
from pathlib import Path
import re
import openpyxl

pdf = PdfFileReader('111JUL19-1.pdf')
excel = openpyxl.load_workbook('excelprueba.xlsx')
sheet1 = excel['Sheet1']



#Extract The page
page0Object = pdf.getPage(0)
#Get Count
# page_count = page0Object.getNumPages()
#Extract the Text
page1text = page0Object.extractText()
# print(page1text)



#Combine de text and save as a Text File
# with Path('pdfToText.txt').open(mode='w') as output_file:
#     text = ''
#     for page in pdf.pages:
#         text += page.extractText()
#         output_file.write(text)

#         #Where is the word

#         word_pages = []
#         Word = 'DOMICILIACIÓN'
#         for page in pdf.pages:
#             page_num = page['/StructParents']
#             page_text = page.extractText()

#             if Word in page_text:
#                 word_pages.append(page_num)


# print(word_pages)

reg: '/\b(Domiciliacion)\b/i'

last_row_number = sheet1.max_row
print(last_row_number)
sheet1.cell(column=1,row=last_row_number + 1) == domiciliacion

excel.save('excelprueba.xlsx')
