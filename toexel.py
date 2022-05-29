


import re
import pdfplumber
# import pandas as pd
from collections import namedtuple
import openpyxl
import sys
import json 

 

excel = openpyxl.load_workbook(filename = './Plantillas modelos impuestos UHY.xlsx')




# nameOfTheFile =s
# nameOfTheFile =  data['name']

# print(modelForm)
# print(nameOfTheFile)




monthRegex = re.compile(r"(0?[1-9]|[1][0-2])$")

valuesRegex =re.compile(r"[0-9]*\.[0-9]+")

# month = ''
lines = []
total_check = 0
cuotas = 0
bases = 0

pattern = re.compile(r"([+-]?(?=\\.\\d|\\d)(?:\\d+)?(?:\\.?\\d*))(?:[eE]([+-]?\\d+))?,([+-]?(?=\\.\\d|\\d)(?:\\d+)?(?:\\.?\\d*))(?:[eE]([+-]?\\d+))?", re.IGNORECASE)

def model111IRPF(file):

   sheet1 = excel['IRPF 111']
   monthRegex = re.compile(r"(0?[1-9]|[1][0-2])$")
   month = ''
   with pdfplumber.open(file) as pdf:
       pages = pdf.pages
       for page in pdf.pages:
           text = page.extract_text()
           for line in text.split('\n') :
               #print(line.split())
               if(line.find('Período') != -1):
                   month = int(monthRegex.search(line).group(1)) + 3
                   items = line.split()
                #    print(month, ' ----------------------')
                #    print(items)


               elif((line.startswith('Rendimientos dinerarios')) & (line.find('01')  != -1)): 
                   items = line.split()
                   perceptores = float(items[4].replace(".","").replace(",","."))
                   percepciones = float(items[6].replace(".","").replace(",","."))
                   retenciones = float(items[8].replace(".","").replace(",","."))
                   sheet1.cell(column= month,row=8, value = perceptores)
                   sheet1.cell(column= month,row=9, value = percepciones)
                   sheet1.cell(column= month,row=10, value = retenciones)
               elif((line.startswith('Rendimientos en especie')) & (line.find('04')  != -1)):
                   #TODO
                   items = line.split()

               elif((line.startswith('Rendimientos dinerarios')) & (line.find('07')  != -1)): 
                   items = line.split()
                #    print(items)
                   perceptores = float(items[4].replace(".","").replace(",","."))
                   percepciones = float(items[6].replace(".","").replace(",","."))
                   retenciones = float(items[8].replace(".","").replace(",","."))
                   sheet1.cell(column= month,row=20, value = perceptores)
                   sheet1.cell(column= month,row=21, value = percepciones)
                   sheet1.cell(column= month,row=22, value = retenciones)
               elif((line.startswith('Rendimientos en especie')) & (line.find('10')  != -1)):
                   print(line)
                   #TODO
                   items = line.split()
               elif((line.startswith('Premios en metálico')) & (line.find('13')  != -1)):
                   #TODO
                   items = line.split()
               elif((line.startswith('Premios en especie')) & (line.find('16')  != -1)):
                   #TODO
                   items = line.split() 
               elif((line.startswith('Percepciones dinerarias')) & (line.find('19')  != -1)):
                   #TODO
                   items = line.split()
               elif((line.startswith('Percepciones en especie')) & (line.find('22')  != -1)):
                   #TODO
                   items = line.split()
               elif((line.startswith('Contraprestaciones dinerarias o en especie')) & (line.find('25')  != -1)):
                   #TODO
                   items = line.split()
               elif((line.startswith('Total liquidación')) & (line.find('25')  != -1)):
                   #TODO
                   items = line.split()
               
               #excel.save('Plantillas modelos impuestos UHY.xlsx') 
   


def model303Iva(file):
   month = ''
   sheet1 = excel['IVA']
   with pdfplumber.open(file) as pdf:
    pages = pdf.pages
    for page in pdf.pages:
        text = page.extract_text()
        for line in text.split('\n'):
            
            if(line.find('Período ') != -1):
                
                month = int(monthRegex.search(line).group(1)) + 4
                # print(month)
                
         
                
            # TIPO 4 %  
            elif((line.find('02') != -1) &( line.find('01') != -1)):
                
                lineEdited = line.replace(".","").replace(",",".")
                # print(lineEdited)

                matches = valuesRegex.findall(lineEdited)   
                if(len(matches) > 1 ):                  
                   bases = float(matches[0])
                   cuotas = float(matches[2]) 
                   sheet1.cell(column= month,row=7, value = bases)
                   sheet1.cell(column= month,row=19, value = cuotas)
              
            # 10%
            elif((line.startswith('Régimen general')) & (line.find('04')  != -1)):
                    
                lineEdited = line.replace(".","").replace(",",".")
                # print(lineEdited)
                matches = valuesRegex.findall(lineEdited)   
                if(len(matches) > 1 ):                 
                    bases = float(matches[0])
                    cuotas = float(matches[2]) 
                    sheet1.cell(column= month,row=8, value = bases)
                    sheet1.cell(column= month,row=20, value = cuotas)
               
             # 21%   
            elif(line.startswith('07')  ):
                lineEdited = line.replace(".","").replace(",",".")
                # print(lineEdited)

                matches = valuesRegex.findall(lineEdited)
                # print(matches[0])  
                # items = line.split()
                if(len(matches) > 1 ):   
                    bases = float(matches[0])               
                    cuotas = float(matches[2])
                    sheet1.cell(column= month,row=9, value = bases)
                    sheet1.cell(column= month,row=21, value = cuotas)
                
            # Adquisiciones intracomunitarias de bienes y servicios.
            elif(line.startswith('Adquisiciones intracomunitarias de bienes y servicios.')  ):
                lineEdited = line.replace(".","").replace(",",".")
                matches = valuesRegex.findall(lineEdited)
                # print(matches) 
                if(len(matches) >= 1 ):                  
                    bases = float(matches[0])  
                    cuotas = float( matches[1])
                    sheet1.cell(column= month,row=22, value = cuotas)
                    sheet1.cell(column= month,row=10, value = bases)
                
            #  ELIF Otras operaciones con inversión del sujeto pasivo'
            elif(line.startswith('Otras operaciones con inversión del sujeto pasivo')  ):
                lineEdited = line.replace(".","").replace(",",".")
                matches = valuesRegex.findall(lineEdited)
                # print(matches)     
                if(len(matches) >= 1 ):              
                    bases = float(matches[0])  
                    cuotas = float( matches[1])
                    sheet1.cell(column= month,row=23, value = cuotas)
                    sheet1.cell(column= month,row=11, value = bases)  
            
            # Modificación bases y cuotas
            elif((line.find('bases y cuotas') != -1 ) & (line.find('14') != -1  )):
             
               lineEdited = line.replace(".","").replace(",",".")
               matches = valuesRegex.findall(lineEdited)             
            
               if(len(matches) >= 1 ):   
                    cuotas = float(matches[1])               
                    bases = float(matches[0])               
                    sheet1.cell(column= month,row=12, value = bases)
                    sheet1.cell(column= month,row=24, value = cuotas)
                
                #  ELIF Recargo equivalencia
            elif((line.find('16') != -1 ) & (line.find('17') != -1 ) & (line.find('18') != -1) ):
                
               lineEdited = line.replace(".","").replace(",",".")
               matches = valuesRegex.findall(lineEdited)             
               print(matches)
               if(len(matches) > 1 ):   
                    cuotas = float(matches[2])               
                    bases = float(matches[0])               
                    sheet1.cell(column= month,row=13, value = bases)
                    sheet1.cell(column= month,row=25, value = cuotas)    
            elif((line.find('19') != -1 ) & (line.find('20') != -1 ) & (line.find('21') != -1) ):
                
               lineEdited = line.replace(".","").replace(",",".")
               matches = valuesRegex.findall(lineEdited)             
               print(matches)
               if(len(matches) > 1 ):   
                    cuotas = float(matches[2])               
                    bases = float(matches[0])               
                    sheet1.cell(column= month,row=14, value = bases)
                    sheet1.cell(column= month,row=26, value = cuotas)    
            elif((line.find('22') != -1 ) & (line.find('23') != -1 ) & (line.find('24') != -1) ):
                
               lineEdited = line.replace(".","").replace(",",".")
               matches = valuesRegex.findall(lineEdited)             
               print(matches)
               if(len(matches) > 1 ):   
                    cuotas = float(matches[2])               
                    bases = float(matches[0])               
                    sheet1.cell(column= month,row=15, value = bases)
                    sheet1.cell(column= month,row=27, value = cuotas)    
            elif((line.find('25') != -1 ) & (line.find('26') != -1 )):
                
               lineEdited = line.replace(".","").replace(",",".")
               matches = valuesRegex.findall(lineEdited)             
               print(matches)
               if(len(matches) > 1 ):   
                    cuotas = float(matches[1])               
                    bases = float(matches[0]) 
                    #No hay casillas en el excel (?)              
                    # sheet1.cell(column= month,row=15, value = bases)
                    # sheet1.cell(column= month,row=27, value = cuotas)    

              #  Modifi caciones bases y cuotas del recargo de equivalencia 
            elif((line.find('28') != -1 ) & (line.find('29') != -1 )):
               lineEdited = line.replace(".","").replace(",",".")
               matches = valuesRegex.findall(lineEdited)             
               
               if(len(matches) > 1 ):   
                    cuotas = float(matches[1])               
                    bases = float(matches[0])     
                    sheet1.cell(column= month,row=31, value = bases)
                    sheet1.cell(column= month,row=41, value = cuotas)
            elif((line.find('30') != -1 ) & (line.find('31') != -1 )):
               lineEdited = line.replace(".","").replace(",",".")
               matches = valuesRegex.findall(lineEdited)             
               
               if(len(matches) > 1 ):   
                    cuotas = float(matches[1])               
                    bases = float(matches[0])  
                    sheet1.cell(column= month,row=32, value = bases)
                    sheet1.cell(column= month,row=42, value = cuotas)      
            #ELIF Por cuotas soportadas en las importaciones de bienes corrientes
            elif((line.find('32') != -1 ) & (line.find('33') != -1 )):
               lineEdited = line.replace(".","").replace(",",".")
               matches = valuesRegex.findall(lineEdited)             
               
               if(len(matches) > 1 ):   
                    cuotas = float(matches[1])               
                    bases = float(matches[0])  
                    sheet1.cell(column= month,row=33, value = bases)
                    sheet1.cell(column= month,row=43, value = cuotas)      
             
            #  Por cuotas soportadas en las importaciones de bienes de inversión
            elif((line.find('34') != -1 ) & (line.find('35') != -1 )):
               lineEdited = line.replace(".","").replace(",",".")
               matches = valuesRegex.findall(lineEdited)             
               
               if(len(matches) > 1 ):   
                    cuotas = float(matches[1])               
                    bases = float(matches[0])  
                    sheet1.cell(column= month,row=34, value = bases)
                    sheet1.cell(column= month,row=44, value = cuotas)    
            elif((line.find('36') != -1 ) & (line.find('37') != -1 )):
                lineEdited = line.replace(".","").replace(",",".")
                matches = valuesRegex.findall(lineEdited)             
               
                if(len(matches) > 1 ):   
                    cuotas = float(matches[1])               
                    bases = float(matches[0])  
                    sheet1.cell(column= month,row=35, value = bases)
                    sheet1.cell(column= month,row=45, value = cuotas)
                
            #  En adquisiciones intracomunitarias de bienes de inversión 
            elif((line.find('38') != -1 ) & (line.find('39') != -1 )):
              lineEdited = line.replace(".","").replace(",",".")
              matches = valuesRegex.findall(lineEdited)             
               
              if(len(matches) > 1 ):   
                    cuotas = float(matches[1])               
                    bases = float(matches[0])  
                    sheet1.cell(column= month,row=36, value = bases)
                    sheet1.cell(column= month,row=46, value = cuotas)
            elif((line.find('40') != -1 ) & (line.find('41') != -1 )):           
              lineEdited = line.replace(".","").replace(",",".")
              matches = valuesRegex.findall(lineEdited)             
               
              if(len(matches) > 1 ):   
                    cuotas = float(matches[1])               
                    bases = float(matches[0])                 
                    sheet1.cell(column= month,row=37, value = bases)
                    sheet1.cell(column= month,row=47, value = cuotas)   
            
         
            # ELIF Compensaciones Régimen Especial A.G. y P.
            elif((line.startswith('Compensaciones Régimen')) & (line.find('42') != -1 )):           
              lineEdited = line.replace(".","").replace(",",".")
              matches = valuesRegex.findall(lineEdited)             
               
              if(len(matches) > 0 ):   
                    cuota = float(matches[0])      
                    sheet1.cell(column= month,row=48, value = cuota)   

            # ELIF Regularización bienes de inversión
            elif((line.startswith('Regularización bienes de inversión')) & (line.find('43') != -1 )):           
              lineEdited = line.replace(".","").replace(",",".")
              matches = valuesRegex.findall(lineEdited)             
               
              if(len(matches) > 0 ):   
                    cuota = float(matches[0])      
                    sheet1.cell(column= month,row=49, value = cuota)  
            # ELIF Regularización por aplicación del porcentaje definitivo de prorrata
            elif((line.startswith('Regularización por aplicación')) & (line.find('44') != -1 )):           
              lineEdited = line.replace(".","").replace(",",".")
              matches = valuesRegex.findall(lineEdited)             
               
              if(len(matches) > 0 ):   
                    cuota = float(matches[0])      
                    sheet1.cell(column= month,row=50, value = cuota) 
            # ELIF Entregas intracomunitarias de bienes y servicios
            

            elif((line.startswith('Entregas intracomunitarias de bienes y servicios')) & (line.find('59') != -1 )):           
              lineEdited = line.replace(".","").replace(",",".")
              matches = valuesRegex.findall(lineEdited)             
               
              if(len(matches) > 0 ):   
                    cuota = float(matches[0])      
                    sheet1.cell(column= month,row=60, value = cuota)  
            
            elif((line.startswith('Exportaciones y operaciones asimiladas')) & (line.find('60') != -1 )):
              lineEdited = line.replace(".","").replace(",",".")
              matches = valuesRegex.findall(lineEdited)             
               
              if(len(matches) > 0 ):   
                    cuota = float(matches[0])                      
                    sheet1.cell(column= month,row=61, value = cuota)
           
            elif((line.find('61') != -1 ) & (line.find('Operaciones') != -1 ) ):
              lineEdited = line.replace(".","").replace(",",".")
              matches = valuesRegex.findall(lineEdited)             
               
              if(len(matches) > 0 ):   
                    cuota = float(matches[0])               
                    sheet1.cell(column= month,row=62, value = cuota)
           
            
            # ELIF Importes de las entregas de bienes y prestaciones de servicios...
            elif((line.find('62') != -1 ) & (line.find('63') != -1 )):           
              lineEdited = line.replace(".","").replace(",",".")
              matches = valuesRegex.findall(lineEdited)             
               
              if(len(matches) > 1 ):   
                    cuotas = float(matches[1])               
                    bases = float(matches[0])
                    #casilla de excel no informada (?)                 
                    # sheet1.cell(column= month,row=37, value = bases)
                    # sheet1.cell(column= month,row=47, value = cuotas)  
            # ELIF Importes de las adquisiciones de bienes y servicios a las que sea de aplicación
            elif((line.find('73') != -1 ) & (line.find('75') != -1 )):           
              lineEdited = line.replace(".","").replace(",",".")
              matches = valuesRegex.findall(lineEdited)             
               
              if(len(matches) > 1 ):   
                    cuotas = float(matches[1])               
                    bases = float(matches[0])
                    #casilla de excel no informada (?)                 
                    # sheet1.cell(column= month,row=37, value = bases)
                    # sheet1.cell(column= month,row=47, value = cuotas)  
            
            
           # RESULTADO
        
           # ELIF Regularización cuotas art. 80.Cinco.5ª LIVA
            elif((line.startswith('Regularización cuotas')) & (line.find('76') != -1 )):
             lineEdited = line.replace(".","").replace(",",".")
             matches = valuesRegex.findall(lineEdited)             
               
             if(len(matches) > 0 ):   
                cuota = float(matches[0])
                    #casilla de excel no informada (?)                        
                    #sheet1.cell(column= month,row=61, value = cuota)  
           #TODO ELIF Suma de resultados ( [46] + [58] + [76] )
            elif((line.startswith('Suma de resultados')) & (line.find('64') != -1 )):
                lineEdited = line.replace(".","").replace(",",".")
                matches = valuesRegex.findall(lineEdited)             
               
                if(len(matches) > 0 ):   
                    cuota = float(matches[0]) 
           
        
           # ELIF Atribuible a la Administración del Estado ( [46] + [58] + [76] )
            elif(line.startswith('Atribuible') & (line.find('66') != -1 ) & (line.find('65') != -1 ) ):
                
                lineEdited = line.replace(".","").replace(",",".")
                matches = valuesRegex.findall(lineEdited)             
              
                if(len(matches) > 0 ):
                    cuota = float(matches[1])                
                    sheet1.cell(column= month,row=57, value = cuota)
            elif(line.startswith('IVA a la importación') & (line.find('77') != -1 )):
                
                lineEdited = line.replace(".","").replace(",",".")
                matches = valuesRegex.findall(lineEdited)             
                
                if(len(matches) > 0 ):
                    cuota = float(matches[0])                
                    sheet1.cell(column= month,row=58, value = cuota)
            elif(line.startswith('Cuotas a compensar') & (line.find('67') != -1 )):
                
                lineEdited = line.replace(".","").replace(",",".")
                matches = valuesRegex.findall(lineEdited)             
                
                if(len(matches) > 0 ):
                    cuota = float(matches[0])                
                    sheet1.cell(column= month,row=59, value = cuota)
                               
        
            # excel.save('Plantillas modelos impuestos UHY.xlsx')



def getDataFromPdfAndSaveExcel():
    data = json.loads(sys.stdin.readline())    
    sys.stdout.flush()
    modelForm = data['type']   
    for name in data['names']:
        print(name)
        file = './uploads/' + name
        if(modelForm == "303"):
           print('pasa 303')
           model303Iva(file)
           
           

        elif(modelForm == "111"):
           print('pasa 111')
           model111IRPF(file)
           
    excel.save('Plantillas modelos impuestos UHY.xlsx')       
    sys.stdout.close()
    
    #excel.save('Plantillas modelos impuestos UHY.xlsx')
           

getDataFromPdfAndSaveExcel()


